# coding=utf-8

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import UnexpectedAlertPresentException
import os
from time import sleep
import argparse
from collections import defaultdict
from bs4 import BeautifulSoup
import lxml
import openpyxl


class Scraper(object):
    """docstring for Scraper"""
    URL = 'https://www.autoscout24.de/'

    def __init__(self, chrome_path, filename, max_price, min_price):
        self.page = 1
        self.row = 1
        self.filename = filename
        self.max_price = max_price
        self.min_price = min_price
        self.brands = ['Mercedes-Benz', 'BMW', 'Volkswagen', 'Audi', 'Ford',
                       'Porsche', 'Opel', 'Citroen', 'Fiat', 'Honda',
                       'Hyundai', 'Mazda', 'Mitsubishi', 'Nissan', 'Peugeot',
                       'Renault', 'Toyota']
        self.chars = ['Getriebeart', 'Kraftstoff', 'Marke', 'Modell',
                      'Erstzulassung', 'Außenfarbe', 'Anzahl Türen',
                      'Sitzplätze']
        self.driver = self.chrome_setup(chrome_path)
        self.wb, self.sheet = self.create_excel(self.filename)

    def chrome_setup(self, path, headless=False):
            """Setting up a chrome driver"""
            chrome_options = webdriver.ChromeOptions()

            if headless:
                user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 \
                (KHTML, like Gecko) Chrome/60.0.3112.50 Safari/537.36'

                chrome_options.add_argument(f'user-agent={user_agent}')
                chrome_options.add_argument('--headless')

            driver = webdriver.Chrome(os.path.abspath(path),
                                      options=chrome_options)
            if headless:
                driver.set_window_size(1440, 900)

            return driver

    # Helper function for checking for the presence of a web element.
    def _is_element_displayed(self, elem_text, elem_type):
        if elem_type == "class":
            try:
                out = self.driver.find_element_by_class_name(
                    elem_text).is_displayed()
            except UnexpectedAlertPresentException:
                self.handle_alert()
                self._is_element_displayed(elem_text, elem_type)
            except (NoSuchElementException, TimeoutException):
                out = False
        elif elem_type == "id":
            try:
                out = self.driver.find_element_by_id(elem_text).is_displayed()
            except (NoSuchElementException, TimeoutException):
                out = False
        else:
            raise ValueError("arg 'elem_type' must be either 'class' or 'id'")

        return out

    # If captcha page is displayed, this function will run indefinitely until
    # the captcha page is no longer displayed (checks for it every 30 seconds).
    # Purpose of the function is to "pause" execution of the scraper until the
    # user has manually completed the captcha requirements.
    def _pause_for_captcha(self):
        while True:
            sleep(30)
            if not self._is_element_displayed("g-recaptcha", "class"):
                break

    def check_for_captcha(self):
        if self._is_element_displayed("g-recaptcha", "class"):
            print("\nCAPTCHA!\n"
                  "Manually complete the captcha requirements.\n"
                  "Once that's done, if the program was in the middle of "
                  "scraping (and is still running), it should resume scraping "
                  "after ~30 seconds.")
            self._pause_for_captcha()

    def handle_alert(self):
        alert = self.driver.switch_to.alert
        alert.accept()

    def create_excel(self, filename):
        """Function to load excel file with data"""
        columns = ['Phone', 'Preis', 'Marke', 'Modell', 'Kilometerstand',
                   'Stadt', 'Getriebeart', 'Kraftstoff', 'Erstzulassung',
                   'Außenfarbe', 'Anzahl Türen', 'Sitzplätze']
        wb = openpyxl.Workbook()
        col = 1
        if self.row == 1:
            for c in columns:
                wb.active.cell(row=self.row, column=col, value=c)
                col += 1
            self.row += 1
            wb.save(filename)

        return wb, wb.active

    def save_data(self, data):
        list_of_columns = [s.value for s in self.sheet[1]]
        for key, value in data.items():
            r = self.sheet.cell(row=self.row,
                                column=list_of_columns.index(key) + 1)
            r.value = ''.join(value)
        self.wb.save(self.filename)

    def start_page(self):
        self.driver.get(self.URL)
        box = self.driver.find_element_by_id('cl-filter-home-car-fragment')
        box.find_element_by_tag_name('button').click()
        sleep(1.5)

    def set_filters(self):
        url = '''https://www.autoscout24.de/lst?priceto=%s&desc=0&size=20&\
custtype=P&page=%d&fc=3&cy=D&ac=0&pricefrom=%s&sort=standard&ustate=N%%2C\
U&atype=C''' % (self.max_price, self.page, self.min_price)
        self.driver.get(url)
        sleep(1.5)

    def get_links(self):
        links = [link.get_attribute('href') for link in
                 self.driver.find_elements_by_xpath(
            "//a[@data-item-name='detail-page-link']")]

        return links

    def get_details(self):
        def get_additional(classnames, data):
            for classname in classnames:
                where_to = soup.find_all('div', {'class': classname})[-1]
                dts = [d.text.strip() for d in where_to.find_all('dt')]
                dds = [d.text.strip() for d in where_to.find_all('dd')]
                for x in range(len(dts)):
                    if (dts[x] in self.chars):
                        data[dts[x]].append(dds[x])

            return data

        data = defaultdict(list)
        soup = BeautifulSoup(self.driver.page_source, 'lxml')

        marke = soup.find(
            'span', {'class': 'cldt-detail-makemodel sc-ellipsis'}).text
        if marke in self.brands:
            return None
        try:
            data['Phone'].append(self.driver.find_element_by_css_selector(
                'a.sc-btn-ghost.cldt-stage-call-btn.sc-ellipsis').text.
                replace('+49 (0)', '').replace('-', ''))
        except NoSuchElementException:
            return None
        try:
            data['Preis'].append(soup.find(
                'div', {'class': 'cldt-price'}).text.replace('€ ', '').replace(
                ',-', '').strip())
        except AttributeError:
            data['Preis'].append('-')
        try:
            km = soup.find(
                'span', {'class': 'sc-font-l cldt-stage-primary-keyfact'}).text
            if 'km' in km:
                data['Kilometerstand'].append(km)
        except AttributeError:
            data['Kilometerstand'].append('-')
        try:
            city = soup.find('div', {'class': 'cldt-stage-vendor-data'}).find(
                'span', {'class': 'sc-font-bold'}).text
            data['Stadt'].append(city)
        except AttributeError:
            data['Stadt'].append('-')

        data = get_additional([
            'cldt-categorized-data cldt-data-section sc-pull-left',
            'cldt-categorized-data cldt-data-section sc-pull-right',
            'cldt-data-section sc-grid-col-s-12'], data)

        return data

    def search(self):
        self.start_page()
        self.set_filters()
        self.check_for_captcha()

        pagination = self.driver.find_element_by_class_name('sc-pagination')
        last_page = pagination.find_elements_by_tag_name('li')[-2].text

        while self.page != int(last_page):
            links = self.get_links()
            for link in links:
                self.driver.get(link)
                self.check_for_captcha()
                sleep(1.5)
                details = self.get_details()

                if details is not None:
                    self.save_data(details)
                    self.row += 1

            self.page += 1
            print(f'Next page is {self.page}')
            self.set_filters()
            self.check_for_captcha()


parser = argparse.ArgumentParser()
parser.add_argument('min_price', help='minimum price for price range')
parser.add_argument('max_price', help='maximum price for price range')
args = parser.parse_args()

min_price = args.min_price
max_price = args.max_price

scraper = Scraper('chromedriver.exe', 'AutoscoutData.xlsx',
                  max_price, min_price)
scraper.search()
